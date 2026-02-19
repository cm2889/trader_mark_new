from multiprocessing import context
import os 
import base64 
from django.shortcuts import render, redirect
from django.core.paginator import Paginator
from django.core.paginator import PageNotAnInteger, EmptyPage
from django.views.generic import ListView, CreateView, UpdateView, DetailView, DeleteView
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.urls import reverse_lazy
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AdminPasswordChangeForm
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.db.models import Q
from django.views.decorators.http import require_GET 
from django.template.loader import render_to_string
from django.core.exceptions import ValidationError
from django.urls import reverse
from django.conf import settings
from django.core.cache import cache
from PIL import Image
from io import BytesIO
from django.utils.text import slugify 
from django.utils import timezone
from datetime import datetime 
from django.db.models import Sum, Count, Q, F
from backend.models import UniformStockTransactionLog
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


from backend.models import (
    WebImages, SiteSettings, LoginLog, UserMenuPermission, 
    BackendMenu, 
    Visitor, 
    Nationality, Employee, Employment, Passport, DrivingLicense, 
    HealthInsurance, Contact, Address, Vehicle, InsuranceClaim, 
    VehicleHandover, TrafficViolation,ViolationType, TrafficViolationPenalty, 
    VehicleInstallment, VehicleMaintenance, VehicleAccident, VehicleAssign, 
    ViolationType, VehicleMaintananceType, 
    Uniform, UniformStock, UniformIssuance, UniformClearance, VehiclePurchase,
    Company
)

from backend.forms import (
    CustomUserLoginForm, NationalityForm, EmployeeForm, EmploymentForm, VehicleMaintananceTypeForm, 
    PassportForm, DrivingLicenseForm, HealthInsuranceForm, ContactForm, 
    AddressForm, UserCreateForm, VehicleForm, VisitorForm, InsuranceClaimForm, 
    VehicleHandoverForm, TrafficViolationForm, ViolationTypeForm, TrafficViolationPenaltyForm, 
    VehicleInstallmentForm, VehicleMaintenanceForm, VehicleAccidentForm, VehicleAssignForm, 
    UniformForm, UniformStockForm, UniformIssuanceForm, UniformClearanceForm,
    VehiclePurchaseForm, InstallmentPaymentForm, CompanyForm
) 

from backend.common_func import checkUserPermission

def paginate_data(request, page_num, data_list):
    items_per_page, max_pages = 10, 10 
    paginator = Paginator(data_list, items_per_page)
    last_page_number = paginator.num_pages 

    try:
        data_list = paginator.page(page_num)
    except PageNotAnInteger:
        data_list = paginator.page(1)
    except EmptyPage:
        data_list = paginator.page(last_page_number)

    current_page = data_list.number
    start_page = max(current_page - int(max_pages / 2), 1)
    end_page = start_page + max_pages

    if end_page > last_page_number:
        end_page = last_page_number + 1
        start_page = max(end_page - max_pages, 1)

    paginator_list = range(start_page, end_page)

    return paginator_list, data_list, last_page_number


def import_excel(request):
    # Get all active companies for the export dropdown
    companies = Company.objects.filter(is_active=True, deleted=False).order_by('name')
    
    if request.method == "POST":
        file = request.FILES.get('excel_file')

        if not file:
            messages.error(request, "Please upload an Excel file.") 
            return redirect('import_export:import_excel')

        # Validate file extension
        if not file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, "Please upload a valid Excel file (.xlsx or .xls)")
            return redirect('import_export:import_excel')

        try:
            # Extract company name from filename (without extension)
            filename = file.name
            company_name = os.path.splitext(filename)[0].strip()
            
            # Get or create company based on filename
            company, company_created = Company.objects.get_or_create(
                name=company_name,
                defaults={
                    'created_by': request.user,
                }
            )
            
            if company_created:
                messages.info(request, f"Created new company: {company_name}")
            else:
                messages.info(request, f"Processing data for company: {company_name}")

            # Read Excel file (process first sheet only)
            df = pd.read_excel(file, sheet_name=0)
            
            # Clean column names - remove extra spaces
            df.columns = df.columns.str.strip()

            success_count = 0
            error_count = 0
            errors = []

            # --- Helper functions (defined once, used throughout) ---
            def safe_str(val):
                """Safely convert value to string, handling NaN and float values.
                Converts float integers like 70037744.0 to '70037744'."""
                if pd.isna(val):
                    return ''
                if isinstance(val, float) and val == int(val):
                    return str(int(val)).strip()
                return str(val).strip()

            def parse_date(date_val):
                """Parse date from Excel with multiple format support."""
                if pd.isna(date_val):
                    return None
                if isinstance(date_val, datetime):
                    return date_val.date()
                try:
                    return pd.to_datetime(date_val).date()
                except:
                    return None

            def parse_expiry_date(date_val):
                """Parse expiry date, fixing Excel 2-digit year misinterpretation.
                Excel maps 2-digit years 30-99 to 1930-1999 instead of 2030-2099.
                Expiry dates before 2000 are clearly wrong, so we add 100 years."""
                d = parse_date(date_val)
                if d and d.year < 2000:
                    d = d.replace(year=d.year + 100)
                return d

            def clean_phone(val):
                """Clean phone number:
                - Strips '.0' float suffix from Excel numeric values
                - Removes all special characters except leading '+'
                - Truncates to 20 chars max
                Returns None if empty."""
                s = safe_str(val)
                if not s:
                    return None
                # Preserve leading +, strip all non-digit chars from the rest
                if s.startswith('+'):
                    s = '+' + re.sub(r'[^\d]', '', s[1:])
                else:
                    s = re.sub(r'[^\d]', '', s)
                return s[:20] if s and s != '+' else None

            for index, row in df.iterrows():
                try:
                    # Skip empty rows
                    if pd.isna(row.get('Qid No')):
                        continue

                    # Get QID No (unique identifier)
                    qid_no = safe_str(row.get('Qid No', ''))

                    # Validate QID No
                    if not qid_no:
                        errors.append(f"Row {index + 2}: Missing QID No")
                        error_count += 1
                        continue

                    # Check if employee with this QID already exists
                    existing_employee = Employee.objects.filter(qid_no=qid_no).first()
                    if existing_employee:
                        # Update existing employee
                        employee = existing_employee
                        
                        # Parse name into first_name and last_name
                        full_name = safe_str(row.get('Name', ''))
                        name_parts = full_name.split(maxsplit=1)
                        first_name = name_parts[0] if len(name_parts) > 0 else ''
                        last_name = name_parts[1] if len(name_parts) > 1 else ''

                        # Parse nationality
                        nationality_name = safe_str(row.get('Nationality', ''))
                        nationality = None
                        if nationality_name:
                            nationality, _ = Nationality.objects.get_or_create(
                                name=nationality_name,
                                defaults={'code': nationality_name[:3].upper(), 'created_by': request.user}
                            )

                        # Parse gender
                        gender_val = safe_str(row.get('Gender', '')).upper()
                        gender = 'M' if gender_val == 'MALE' else 'F' if gender_val == 'FEMALE' else 'M'

                        joining_date = parse_date(row.get('Joing Date'))

                        # Update employee fields
                        employee.company = company
                        employee.first_name = first_name
                        employee.last_name = last_name
                        employee.nationality = nationality
                        employee.gender = gender
                        if joining_date:
                            employee.joining_date = joining_date
                        employee.remarks = safe_str(row.get('Remarks', '')) or None
                        employee.updated_by = request.user
                        employee.save()
                        
                    else:
                        # Create new employee (HR File No will be auto-generated)
                        
                        # Parse name into first_name and last_name
                        full_name = safe_str(row.get('Name', ''))
                        name_parts = full_name.split(maxsplit=1)
                        first_name = name_parts[0] if len(name_parts) > 0 else ''
                        last_name = name_parts[1] if len(name_parts) > 1 else ''

                        # Parse nationality
                        nationality_name = safe_str(row.get('Nationality', ''))
                        nationality = None
                        if nationality_name:
                            nationality, _ = Nationality.objects.get_or_create(
                                name=nationality_name,
                                defaults={'code': nationality_name[:3].upper(), 'created_by': request.user}
                            )

                        # Parse gender
                        gender_val = str(row.get('Gender', '')).strip().upper()
                        gender = 'M' if gender_val == 'MALE' else 'F' if gender_val == 'FEMALE' else 'M'

                        joining_date = parse_date(row.get('Joing Date'))

                        # Create new employee (don't set hr_file_no, let it auto-generate)
                        employee = Employee.objects.create(
                            qid_no=qid_no,
                            company=company,
                            first_name=first_name,
                            last_name=last_name,
                            nationality=nationality,
                            gender=gender,
                            joining_at=joining_date or timezone.now(),
                            remarks=safe_str(row.get('Remarks', '')) or None,
                            created_by=request.user,
                        )

                    # Now process related data for this employee (use parse_expiry_date for expiry fields)
                    rp_exp_date = parse_expiry_date(row.get('RP Exp date'))
                    passport_expiry = parse_expiry_date(row.get('Passport Exipry'))
                    license_expiry = parse_expiry_date(row.get('Driving license Expiry'))
                    istemara_expiry = parse_expiry_date(row.get('Istemara Expiry'))

                    # Create or update Employment
                    work_permit_no = safe_str(row.get('Work Permit NO', ''))
                    work_id = safe_str(row.get('Work ID', ''))
                    
                    if work_permit_no and work_id and rp_exp_date:
                        # Parse work status
                        work_sts = safe_str(row.get('Work Sts', '')).upper()
                        work_status_map = {
                            'TALABAT': 'ACTIVE',
                            'SAK': 'ACTIVE',
                            'ACTIVE': 'ACTIVE',
                            'INACTIVE': 'INACTIVE',
                            'ON LEAVE': 'ON_LEAVE',
                            'TERMINATED': 'TERMINATED',
                        }
                        work_status = work_status_map.get(work_sts, 'ACTIVE')

                        # Parse QID renew/lost status
                        qid_renew = safe_str(row.get('QID Renew', '')).upper()
                        qid_lost = safe_str(row.get('QID Lost', '')).upper()

                        employment, _ = Employment.objects.update_or_create(
                            employee=employee,
                            defaults={
                                'rp_expiry_date': rp_exp_date,
                                'work_permit_no': work_permit_no,
                                'work_id': work_id,
                                'work_status': work_status,
                                'qid_renew_status': 'RENEWED' if qid_renew == 'YES' else 'NOT_DUE',
                                'qid_lost_status': 'YES' if qid_lost == 'YES' else 'NO',
                                'created_by': request.user,
                                'updated_by': request.user,
                            }
                        )

                    # Create or update Passport
                    passport_no = safe_str(row.get('Passport No', ''))
                    if passport_no and passport_expiry:
                        passport_renew = safe_str(row.get('Passport Renew', '')).upper()
                        Passport.objects.update_or_create(
                            employee=employee,
                            defaults={
                                'passport_no': passport_no,
                                'passport_expiry_date': passport_expiry,
                                'passport_renewed': passport_renew == 'YES',
                                'created_by': request.user,
                                'updated_by': request.user,
                            }
                        )

                    # Create or update Driving License
                    license_no = safe_str(row.get('Driving License', ''))
                    if license_no and license_expiry:
                        license_renew = safe_str(row.get('Driving License Renew', '')).upper()
                        DrivingLicense.objects.update_or_create(
                            employee=employee,
                            defaults={
                                'license_no': license_no,
                                'license_expiry_date': license_expiry,
                                'license_renewed': license_renew == 'YES',
                                'license_renew_status': 'YES' if license_renew == 'YES' else 'NO',
                                'created_by': request.user,
                                'updated_by': request.user,
                            }
                        )

                    # Create or update Health Insurance
                    hamad_health = safe_str(row.get('Hamad Health Card', '')).upper()
                    wm_insurance = safe_str(row.get('WM Insurance', '')).upper()
                    fhc = safe_str(row.get('FHC', '')).upper()

                    if hamad_health or wm_insurance or fhc:
                        HealthInsurance.objects.update_or_create(
                            employee=employee,
                            defaults={
                                'hamad_health_card': hamad_health == 'YES',
                                'wm_insurance': 'YES' if wm_insurance == 'YES' else 'NO',
                                'family_health_card': 'YES' if fhc == 'YES' else 'NO',
                                'created_by': request.user,
                                'updated_by': request.user,
                            }
                        )

                    # Create or update Contact (clean_phone strips '.0', special chars, keeps leading '+')
                    phone_no = clean_phone(row.get('Phone No', ''))
                    if phone_no:
                        Contact.objects.update_or_create(
                            employee=employee,
                            defaults={
                                'phone_no': phone_no,
                                'phone_no_alt': clean_phone(row.get('Phone No 01', '')),
                                'roommate_phone': clean_phone(row.get('Friends/ Room Mate Phone no', '')),
                                'relative_qatar_phone': clean_phone(row.get('Relative Qatar Phone No', '')),
                                'home_phone': clean_phone(row.get('Home Phone No', '')),
                                'home_phone_alt': clean_phone(row.get('Home Phone No 01', '')),
                                'home_email': safe_str(row.get('Home Email', '')) or None,
                                'created_by': request.user,
                                'updated_by': request.user,
                            }
                        )

                    # Create or update Address
                    national_address = safe_str(row.get('N. Address', '')).upper()
                    room_address = safe_str(row.get('Room Address', ''))

                    if national_address or room_address:
                        Address.objects.update_or_create(
                            employee=employee,
                            defaults={
                                'national_address': national_address if national_address == 'YES' else None,
                                'room_address': room_address or None,
                                'created_by': request.user,
                                'updated_by': request.user,
                            }
                        )

                    # Create or update Vehicle (if applicable)
                    plate_no = safe_str(row.get('Bike / Car No', ''))
                    if plate_no and plate_no != '*' and istemara_expiry:
                        vehicle_type = 'BIKE'  # Default, can be enhanced
                        Vehicle.objects.update_or_create(
                            plate_no=plate_no,
                            defaults={
                                'vehicle_type': vehicle_type,
                                'istemara_expiry_date': istemara_expiry,
                                'insurance_name': 'Default Insurance',
                                'insurance_expiry_date': istemara_expiry,
                                'ownership': 'DRIVER',
                                'created_by': request.user,
                                'updated_by': request.user,
                            }
                        )

                    success_count += 1

                except Exception as e:
                    error_count += 1
                    errors.append(f"Row {index + 2}: {str(e)}")
                    continue

            # Display summary messages
            if success_count > 0:
                messages.success(request, f"Successfully imported {success_count} employee records for {company_name}!")
            
            if error_count > 0:
                messages.warning(request, f"Failed to import {error_count} records.")
                for error in errors[:10]:  # Show first 10 errors
                    messages.error(request, error)
                if len(errors) > 10:
                    messages.error(request, f"...and {len(errors) - 10} more errors")

        except Exception as e:
            messages.error(request, f"Error processing Excel file: {str(e)}") 
            return redirect('import_export:import_excel')

    context = {
        'companies': companies
    }
    return render(request, "export_center/import_excel.html", context)


def export_excel(request):
    """
    Export employee data to Excel file with all related information.
    Generates a single sheet with data for a specific company.
    """
    
    try:
        # Get company_id from request (GET or POST)
        company_id = request.GET.get('company_id') or request.POST.get('company_id')
        
        if not company_id:
            messages.error(request, "Please select a company to export.")
            return redirect('import_export:import_excel')
        
        # Get the company
        try:
            company = Company.objects.get(id=company_id, is_active=True)
        except Company.DoesNotExist:
            messages.error(request, "Company not found.")
            return redirect('import_export:import_excel')
        
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Employees"

        # Define headers (matching import format)
        headers = [
            'HR File No', 'Qid No', 'Name', 'Nationality', 'Gender', 'RP Exp date', 
            'Remarks', 'N. Address', 'Joing Date', 'Work Sts', 'Work Permit NO', 
            'Work ID', 'QID Renew', 'QID Lost', 'Passport No', 'Passport Exipry', 
            'Passport Renew', 'Driving License', 'Driving license Expiry', 
            'Driving License Renew', 'Hamad Health Card', 'WM Insurance', 'FHC', 
            'Phone No', 'Phone No 01', 'Friends/ Room Mate Phone no', 
            'Relative Qatar Phone No', 'Room Address', 'Home Phone No', 
            'Home Phone No 01', 'Home Email', 'Bike / Car No', 'Istemara Expiry'
        ]

        # Style definitions
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Helper function to format dates
        def format_date(date_obj):
            if date_obj:
                return date_obj.strftime('%m/%d/%Y')
            return ''

        # Helper function to convert boolean to Yes/No
        def bool_to_yn(val):
            return 'Yes' if val else 'No'

        # Fetch employees for this company
        employees = Employee.objects.filter(
            company=company,
            deleted=False
        ).select_related(
            'nationality'
        ).prefetch_related(
            'employments',
            'passports',
            'driving_licenses',
            'health_insurance',
            'contact',
            'address'
        ).order_by('hr_file_no')

        # Add headers to sheet
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Populate data rows for this company
        row_num = 2
        for employee in employees:
            # Get related objects (handle missing data gracefully)
            employment = employee.employments.filter(deleted=False).first()
            passport = employee.passports.filter(deleted=False).first()
            license = employee.driving_licenses.filter(deleted=False).first()
            health = getattr(employee, 'health_insurance', None)
            contact = getattr(employee, 'contact', None)
            address = getattr(employee, 'address', None)
            # Get vehicle assigned to this employee
            vehicle_assign = VehicleAssign.objects.filter(
                employee=employee,
                is_active=True,
                deleted=False
            ).select_related('vehicle').first()
            vehicle = vehicle_assign.vehicle if vehicle_assign else None

            # Prepare row data
            row_data = [
                employee.hr_file_no,
                employee.qid_no,
                employee.full_name,
                employee.nationality.name if employee.nationality else '',
                'Male' if employee.gender == 'M' else 'Female',
                format_date(employment.rp_expiry_date) if employment else '',
                employee.remarks or '',
                'YES' if address and address.national_address else 'NO',
                format_date(employee.joining_at.date() if employee.joining_at else None),
                employment.get_work_status_display() if employment else '',
                employment.work_permit_no if employment else '',
                employment.work_id if employment else '',
                'YES' if employment and employment.qid_renew_status == 'RENEWED' else 'NO',
                employment.qid_lost_status if employment else 'NO',
                passport.passport_no if passport else '',
                format_date(passport.passport_expiry_date) if passport else '',
                bool_to_yn(passport.passport_renewed) if passport else 'No',
                license.license_no if license else '',
                format_date(license.license_expiry_date) if license else '',
                license.license_renew_status if license else 'NO',
                bool_to_yn(health.hamad_health_card) if health else 'No',
                health.wm_insurance if health else 'No',
                health.family_health_card if health else 'No',
                contact.phone_no if contact else '',
                contact.phone_no_alt if contact else '',
                contact.roommate_phone if contact else '',
                contact.relative_qatar_phone if contact else '',
                address.room_address if address else '',
                contact.home_phone if contact else '',
                contact.home_phone_alt if contact else '',
                contact.home_email if contact else '',
                vehicle.plate_no if vehicle else '',
                format_date(vehicle.istemara_expiry_date) if vehicle else '',
            ]

            # Write row data
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

            row_num += 1

        # Auto-adjust column widths for this sheet
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = len(headers[col_num - 1])
            ws.column_dimensions[column_letter].width = min(max_length + 5, 40)

        # Set row height for header
        ws.row_dimensions[1].height = 30

        # Prepare HTTP response with company name as filename
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        # Use company name as filename
        safe_company_name = company.name.replace(' ', '_')
        filename = f'{safe_company_name}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Save workbook to response
        wb.save(response)
        
        return response

    except Exception as e:
        messages.error(request, f"Error exporting data: {str(e)}")
        return redirect('import_export:import_excel')


 